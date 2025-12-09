#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Конвертирует CSV файл в настоящий XLSX формат с правильными типами данных:
- Timestamp -> DateTime
- Числа -> Number
- Логические -> Boolean (ИСТИНА/ЛОЖЬ)
- Строки -> Text
"""

import sys
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def parse_csv_line(line):
    """Парсит CSV строку с учетом кавычек"""
    reader = csv.reader([line])
    try:
        return next(reader)
    except:
        return line.split(',')

def timestamp_to_datetime(ts_str):
    """Конвертирует Unix timestamp в datetime объект"""
    try:
        ts = int(ts_str)
        return datetime.fromtimestamp(ts)
    except (ValueError, TypeError):
        return None

def is_boolean(value):
    """Проверяет, является ли значение логическим (ИСТИНА/ЛОЖЬ)"""
    return value.strip() in ['ИСТИНА', 'ЛОЖЬ']

def boolean_to_bool(value):
    """Конвертирует ИСТИНА/ЛОЖЬ в True/False"""
    return value.strip() == 'ИСТИНА'

def is_number(value):
    """Проверяет, является ли значение числом"""
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False

def convert_to_number(value):
    """Конвертирует строку в число"""
    try:
        if '.' in value:
            return float(value)
        else:
            return int(value)
    except (ValueError, TypeError):
        return value

def csv_to_xlsx(csv_path, xlsx_path):
    """Конвертирует CSV в XLSX с правильными типами данных"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Telemetry Data"
    
    # Стили
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    number_alignment = Alignment(horizontal="right")
    center_alignment = Alignment(horizontal="center")
    
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            
            # Читаем заголовок
            header = next(reader)
            for col_idx, col_name in enumerate(header, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Читаем данные
            row_num = 2
            for row in reader:
                if not row:
                    continue
                    
                for col_idx, value in enumerate(row, 1):
                    if col_idx > len(header):
                        break
                    
                    col_name = header[col_idx - 1] if col_idx <= len(header) else ''
                    cell = ws.cell(row=row_num, column=col_idx)
                    
                    # Обработка по типу колонки
                    if col_name == 'recorded_at':
                        # Timestamp -> DateTime
                        dt = timestamp_to_datetime(value)
                        if dt:
                            cell.value = dt
                            cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                        else:
                            cell.value = value
                    elif col_name in ['flag_A', 'flag_B']:
                        # Логические значения
                        if is_boolean(value):
                            cell.value = boolean_to_bool(value)
                            cell.alignment = center_alignment
                            if boolean_to_bool(value):
                                cell.font = Font(color="4CAF50", bold=True)
                            else:
                                cell.font = Font(color="F44336", bold=True)
                        else:
                            cell.value = value
                    elif col_name in ['voltage', 'temp', 'count']:
                        # Числа
                        if is_number(value):
                            cell.value = convert_to_number(value)
                            cell.alignment = number_alignment
                            if col_name in ['voltage', 'temp']:
                                cell.number_format = '0.00'
                        else:
                            cell.value = value
                    else:
                        # Строки
                        cell.value = value
                
                row_num += 1
            
            # Автоматическая ширина колонок
            for col_idx in range(1, len(header) + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                for row in ws[col_letter]:
                    try:
                        if len(str(row.value)) > max_length:
                            max_length = len(str(row.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width
        
        wb.save(xlsx_path)
        return True
    except Exception as e:
        print(f"Error converting CSV to XLSX: {e}", file=sys.stderr)
        return False

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: csv_to_xlsx.py <input.csv> <output.xlsx>", file=sys.stderr)
        sys.exit(1)
    
    csv_path = sys.argv[1]
    xlsx_path = sys.argv[2]
    
    if csv_to_xlsx(csv_path, xlsx_path):
        print(f"Successfully converted {csv_path} to {xlsx_path}")
        sys.exit(0)
    else:
        sys.exit(1)

